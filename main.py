import bs4
import requests
from bs4 import BeautifulSoup
import pandas as pd
import openpyxl
import re
from openpyxl import load_workbook

def find_codechef_score(username):
  url = "https://codechef.com/users/" + str(username)
  response = requests.get(url)
  # Check if the request was successful
  if response.status_code == 200:
    # Parse the HTML content of the web page
    parsed_content = bs4.BeautifulSoup(response.content, 'html.parser')
    all_tags=parsed_content.find_all('h3')
    
    total_problems=0
    c=0
    for each_tag in all_tags:
        pattern = r'\((\d+)\)'
        match = re.search(pattern, each_tag.text)
        if match:
            c=c+1
            if c==3:
              continue
            number = match.group(1)
            total_problems=total_problems+int(number)
    return total_problems
  else:
    return 0

def find_leetcode_score(username):
    username = str(username)
    query = """
    {
      matchedUser(username: "%s") {
        submitStats: submitStatsGlobal {
          acSubmissionNum {
            difficulty
            count
          }
        }
      }
    }
    """ % username
    url = 'https://leetcode.com/graphql'
    response = requests.post(url, json={'query': query})
    if response.status_code == 200:
        try:
            data = response.json()
            if 'data' in data and 'matchedUser' in data['data'] and 'submitStats' in data['data']['matchedUser']:
                ac_submission_nums = data['data']['matchedUser']['submitStats']['acSubmissionNum']
                if 'count' in ac_submission_nums[0]:
                  total_problems = ac_submission_nums[0]['count']
                return total_problems
            else:
                return 0
        except Exception as e:
            return 0
    else:
        return 0

raw_data=pd.read_csv('profiles_data.csv')
names = raw_data['Names'].tolist()
codechef_usernames=raw_data['CodechefProfile'].tolist()
codechef_scores=[]
leetcode_usernames=raw_data['LeetcodeProfile'].tolist()
leetcode_scores=[]
total_score=[]
data=[]
for user in codechef_usernames:
  codechef_scores.append(find_codechef_score(user))
for user in leetcode_usernames:
  leetcode_scores.append(find_leetcode_score(user))
for i in range(len(leetcode_scores)):
  total_score.append(leetcode_scores[i]*50+codechef_scores[i]*20)
for i in range(len(leetcode_usernames)):
  new_row=[names[i],leetcode_scores[i],codechef_scores[i],total_score[i]]
  data.append(new_row)
roll_nos=raw_data['Roll_No'].tolist()
d={
  'Roll Number':roll_nos,
  'Name':names,
  'Leetcode':leetcode_scores,
  'CodeChef':codechef_scores,
  'Total Score':total_score
}
df = pd.DataFrame(d)
df=df.sort_values(by='Total Score',ascending=False)
file_path='scores.xlsx'
excel_file=openpyxl.load_workbook(file_path)
ws=excel_file.active
for row in ws.iter_rows():
  for cell in row:
    cell.value=''
excel_file.save(file_path)
df.to_excel(file_path,index=False)

# Load the workbook
workbook = load_workbook('scores.xlsx')

# Select the active sheet
sheet = workbook.active

# Define the header for the new column
new_header = "    Rank"

# Insert the new column at the beginning
sheet.insert_cols(1)

# Set the header for the new column
sheet.cell(row=1, column=1, value=new_header)

# Populate the new column with ranks starting from 1
max_row = sheet.max_row
for row_num in range(2, max_row + 1):
    sheet.cell(row=row_num, column=1, value=row_num - 1)

# Save the modified workbook
workbook.save('scores.xlsx')

