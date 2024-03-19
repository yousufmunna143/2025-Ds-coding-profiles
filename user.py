import streamlit as st
from openpyxl import load_workbook
import pandas as pd

st.set_page_config(page_title="2025 DS Leader board",
                   layout="wide")
st.title("2025 DS Batch LeaderBoard")
workbook = load_workbook('scores.xlsx')

# Select the active sheet
sheet = workbook.active

# Define the header for the new column
new_header = "Rank"

# Insert the new column at the beginning
sheet.insert_cols(1)

# Set the header for the new column
sheet.cell(row=1, column=1, value=new_header)

# Populate the new column with ranks starting from 1
max_row = sheet.max_row
for row_num in range(2, max_row + 1):
    sheet.cell(row=row_num, column=1, value=row_num - 1)

# Save the modified workbook
workbook.save('scores.xlsx')

data=pd.read_excel('scores.xlsx')
df=pd.DataFrame(data)
st.dataframe(df, use_container_width=True, hide_index=True)
